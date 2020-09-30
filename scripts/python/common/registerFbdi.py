import sys
import re
import time
import array
import cx_Oracle
from openpyxl import load_workbook

def registerColumns(tab):
    wb.active = tab
    ws = wb.active
    row_count = ws.max_row
    column_count = ws.max_column
    col_order = 1
    mandatory = "N"
    arr = []
    tab_col_name = None
    data_type = None
    help = None

    for x in range(50):
        val=ws.cell(row=x+1,column=1).internal_value
        if val != None and "Required" in val:
            heading = x+2
            break

    for y in range(column_count):
        colName=ws.cell(row=heading, column=y+1).internal_value
        if colName == None:
            break
        else:
	    if "*" in colName:
		mandatory = "Y"
	    else:
		mandatory = "N"
	
            colName = colName.replace("*","")
            colName = colName.strip()
            colName = re.sub('\s+',' ',colName)
            colName = re.sub(re.compile(r'\s+'), '_', colName)
            colName = colName.replace("-","_")
            try:
                cmnt = ws.cell(row=heading, column=y+1).comment.text
                del arr
                arr = []
                for line in cmnt.splitlines():
                    if line != '':
                        arr.append(line)
                try:
                    tab_col_name = arr[0]
                    data_type = arr[1]
                    help = arr[2]
                except IndexError:
                    None
                cur.callproc('xxcm_metadata_pub.insert_fbdi_columns', (col_order, colName, tab_col_name, data_type, help, None, mandatory))
            except AttributeError as error:
                cur.callproc('xxcm_metadata_pub.insert_fbdi_columns', (col_order, colName, None, None, None, None, mandatory))
	    col_order = col_order + 1

file = sys.argv[1]
metaData =  sys.argv[2]

# Get connection
con = cx_Oracle.connect("cms", "oracle", "localhost:1527/xepdb1")
cur = con.cursor()

wb = load_workbook(file)
sheet_count=0
	
for sheet in wb:
    if sheet_count > 0:
        try:
            cur.callproc('xxcm_metadata_pub.insert_fbdi_record', (metaData, sheet.title))
        except cx_Oracle.DatabaseError, exception:
            printException (exception)
            exit (1)
        registerColumns(sheet_count)
    sheet_count = sheet_count + 1
